import openpyxl
import sys

# Get the input file path from the command line argument
input_file = sys.argv[1]

# Open the input file
with open(input_file, 'r') as f:
    data = f.read()

# Split the data into entries
entries = data.split('ER\n')

# Create a new workbook and select the active worksheet
wb = openpyxl.Workbook()
ws = wb.active

# Write the headers to the worksheet
headers = ['PT', 'PN', 'TI', 'AU', 'AE', 'GA', 'AB', 'TF', 'DC', 'MC', 'IP', 'PD', 'AD', 'PI', 'CP', 'UT']
ws.append(headers)

# Parse each entry and write it to the worksheet
for entry in entries:
    attributes = entry.split('\n')
    row = []
    for header in headers:
        value = ''
        for attribute in attributes:
            if attribute.startswith(header):
                value += attribute.replace(header, '').strip()
        row.append(value)
    ws.append(row)

# Print the number of rows and columns in the worksheet
print(f'Total Rows: {ws.max_row}, Total Columns: {ws.max_column}')

# Get the output file path from the command line argument
output_file = input_file.replace('.txt', '.xlsx')

# Save the workbook to an Excel file
wb.save(output_file)

# Print confirmation message
print(f'Successfully converted {input_file} to {output_file}.')